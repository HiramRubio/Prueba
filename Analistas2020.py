# -*- coding: utf-8 -*-
"""
Created on Thu Jan  7 17:46:05 2021

@author: HRV
"""

from openpyxl import load_workbook
from openpyxl import Workbook


#Variables
Steven = 0
Cris = 0
Alejandra = 0
Victor = 0
Grace = 0


#Funciones
def DetAnalista(value):
    
    global Steven, Cris, Alejandra, Victor, Grace
    Sismo = False
    if(value=='A'): Alejandra = Alejandra +1; Sismo = True
    if(value=='C'): Cris = Cris +1          ; Sismo = True
    if(value=='G'): Grace = Grace +1        ; Sismo = True
    if(value=='V'): Victor = Victor +1      ; Sismo = True
    if(value=='S'): Steven = Steven +1      ; Sismo = True
    return Sismo
    


Archivos = ['enero','febrero','marzo','abril','mayo','junio','julio','agosto','septiembre','octubre','noviembre','diciembre']
#Eventos sistema automatico
Auto = 0
#Cargando cada Archivo
for archivo in Archivos:
    wb = load_workbook(filename = 'Eventos_'+archivo+'_2020.xlsx', data_only =True)
    print(f"Minando mes de: {archivo}")
    #Guardamos todos los nombres de hojas
    sheets = wb.sheetnames
    
    #Recorremos todas las hojas del documento. 
    for name in sheets:    
        sheet = wb[name]
    
        #E6,F6  H6,I6  O6,P6  R6,S6
        #Ciclo 1
        for element in sheet.iter_rows(min_row=6,max_row = 40,min_col=5,max_col=5,values_only=True):
            #Buscamos el nombre de cada analista
            for value in element:   x = DetAnalista(value)
            if(x): Auto = Auto+1
        #Ciclo 2
        for element in sheet.iter_rows(min_row=6,max_row = 40,min_col=8,max_col=8,values_only=True):
            #Buscamos el nombre de cada analista
            for value in element: DetAnalista(value)
        #Ciclo 3
        for element in sheet.iter_rows(min_row=6,max_row = 40,min_col=15,max_col=15,values_only=True):
            #Buscamos el nombre de cada analista
            for value in element: x = DetAnalista(value)
            if(x): Auto = Auto+1
        #Ciclo 4
        for element in sheet.iter_rows(min_row=6,max_row = 40,min_col=18,max_col=18,values_only=True):
            #Buscamos el nombre de cada analista
            for value in element: DetAnalista(value)

Total = Steven+Cris+Alejandra+Victor+Grace
#Resumen
print(f"Sismos 2020!\n Cris: {Cris}\n Alejandra: {Alejandra}\n Grace: {Grace}\n Victor: {Victor}\n Steven: {Steven}")
print(f"Total: {Total} \n Total Automatico: {Auto}")