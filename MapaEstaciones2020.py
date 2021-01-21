# -*- coding: utf-8 -*-
"""
Created on Thu Jan 14 03:05:25 2021

@author: Steven Rubio
Mapa con las estaciones en la red nacional/metropolitana
"""

from openpyxl import load_workbook
from mpl_toolkits.basemap import Basemap
import matplotlib.pyplot as plt
from mpl_toolkits.axes_grid1.inset_locator import zoomed_inset_axes
from mpl_toolkits.axes_grid1.inset_locator import mark_inset

wb = load_workbook(filename = 'Data/CoordenadasAcelerómetros.xlsx', data_only =True)
#Guardamos todos los nombres de hojas
sheets = wb.sheetnames
#Variables
Dep =   ['GUAT','PETEN','IZABA','ZACAP','ESC']
LatD =  [14.74072, 17.12258, 15.52778, 15.07222, 14.0009]
LonD =  [-90.51327,-89.89941,-88.7944,-89.53056,-90.88581]
#Mapa Ciudad
Ciudad = wb['ETNA2 INSTALADOS CIUDAD']
Nacional = wb['ETNA2 INSTALADOS DEP']
#Variables 
Names = []
Lat = []
Lon = []
 
Names2 = []
Lat2 = []
Lon2 = []
            
if(True):
    #Ciclo 1
    for element in Nacional.iter_rows(min_row=2,max_row = 40,min_col=2,max_col=6,values_only=True):
        if(element[0] != None):
            #print(element)
            Names.append(element[0])
            Lat.append(element[3])
            Lon.append(element[4])
    
        #Se obtienen coordenadas máximas
    x1,x2 = max(Lon),min(Lon)
    y1,y2 = max(Lat),min(Lat)
    y2=y2-0.3
    y1=y1+3.5
    x1=x1+1.5
    x2=x2-0.2
    
    #Inicio de plot
    fig = plt.figure(figsize=(12, 8))
    ax = fig.add_subplot(111)
    
    #Primer Basemap
    m = Basemap(resolution='i', # c, l, i, h, f or None
            projection='tmerc', 
            lat_0=15.74, lon_0=-90.15,
            llcrnrlon=x2, llcrnrlat=y2,urcrnrlon=x1, urcrnrlat=y1, epsg=4326)
    
    
 
    #Ciclo 2
    for element in Ciudad.iter_rows(min_row=2,max_row = 40,min_col=2,max_col=6,values_only=True):
        if(element[0] != None):
            #print(element)
            Names2.append(element[0])
            Lat2.append(element[3])
            Lon2.append(element[4])
    
    #Se obtienen coordenadas máximas
    x1,x2 = max(Lon2),min(Lon2)
    y1,y2 = max(Lat2),min(Lat2)
    y2=y2-0.3
    y1=y1+0.3
    x1=x1+0.25
    x2=x2-0.2
    
    m.arcgisimage(service='World_Shaded_Relief', xpixels=900, dpi=100,verbose= True)                           
    #m.fillcontinents(color='#f2f2f2',lake_color='#46bcec')
    #m.drawcoastlines(color='#99ebff')
    #Leemos nuestra shapefile
    #Delimitaciones de Guatemala
    m.readshapefile('Data/gtm/gtm_admbnda_adm0_ocha_conred_20190207', 'ej0',linewidth=0.5)
    m.readshapefile('Data/gtm/gtm_admbnda_adm1_ocha_conred_20190207', 'ej1',drawbounds=True,linewidth=0.4)
    
    #Se plotean las estacion
    for lat,lon,name in zip(Lat2,Lon2,Names2):
        x,y = m(lon,lat)
        plot = ax.plot(x,y,'s',marker='^',color='#458403', markersize=5,markeredgecolor = 'k')
        
    #Se plotean las estacion
    for lat,lon,name in zip(Lat,Lon,Names):
        x,y = m(lon,lat)
        plot = ax.plot(x,y,'s',marker='^',color='r', markersize=8,markeredgecolor = 'k')
        
    #Nombre de Departamentos
    """
    for lat,lon,name in zip(LatD,LonD,Dep):
        #Plot del nombre de algunas estaciones
        x,y = m(lon,lat)
        ax.text(x,y,name,fontsize=8,color='k',rotation=0)
        """
if(True):
    
    #Zoomed Map
    axins = zoomed_inset_axes(ax, 6, loc=1)
    axins.set_xlim(x1, x2)
    axins.set_ylim(y1, y2)
    #Inicio de plot
    plt.xticks(visible=False)
    plt.yticks(visible=False)
    
    #Segundo Mapa
    #y2=y2-0.3
    #y1=y1+0.3
    #x1=x1+0.25
    #x2=x2-0.2
    m2 = Basemap(resolution='i', # c, l, i, h, f or None
            projection='tmerc', 
            lat_0=15.74, lon_0=-90.15,
            llcrnrlon=x2+0.15, llcrnrlat=y2+0.25,urcrnrlon=x1-0.20, urcrnrlat=y1-0.25, epsg=4326, ax=axins)
    
    m2.arcgisimage(service='World_Shaded_Relief', xpixels=900, dpi=100,verbose= True)                            
    #m2.fillcontinents(color='#f2f2f2',lake_color='#46bcec')
    #Leemos nuestra shapefile
    #Delimitaciones de Guatemala
    #m2.readshapefile('Data/gtm/gtm_admbnda_adm2_ocha_conred_20190207', 'ej0')
    m2.readshapefile('Data/gtm/gtm_admbnda_adm1_ocha_conred_20190207', 'ej1',drawbounds=True)
    
    #Se plotean las estacon
    for lat,lon,name in zip(Lat2,Lon2,Names2):
        x,y = m(lon,lat)
        plot = axins.plot(x,y,'s',marker='^',color='#458403', markersize=6,markeredgecolor = 'k')
        #Plot del nombre de algunas estaciones
        #if name in CIUDAD:
            #axins.text(x+0.001,y+0.001,name,fontsize=8,color='r',rotation=45,bbox=dict(boxstyle = "square",
          #facecolor = "white"))
    #Insert       
    mark_inset(ax, axins, loc1=2, loc2=4, fc="none", ec="0.5")